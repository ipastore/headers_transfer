"""Unit tests for xlsx_manager — append and reset functionality."""

import os
import tempfile
import unittest

import openpyxl

from xlsx_manager import HEADERS, append_row, get_or_create_wb, reset_rows

PLAYER_1 = {
    "Name": "Matías Abaldo",
    "Birth Date (dd/mm/yyyy)": "03/04/2004",
    "Height (cm)": 172,
    "Weight (kg)": 61,
    "Preferred Foot": "Right",
    "Nationality (ISO2)": "UY",
    "Second Nationality (ISO2)": "DE",
    "Position": "Attacking Midfielder",
    "Second Position": "Right Winger",
    "Club": "Independiente",
    "On loan": True,
    "Club of origin of the Loan": "Defensor Sporting",
    "Representative": "Edgardo Lasalvia",
}

PLAYER_2 = {
    "Name": "Lucas Fernández",
    "Birth Date (dd/mm/yyyy)": "15/07/2001",
    "Height (cm)": 181,
    "Weight (kg)": 75,
    "Preferred Foot": "Left",
    "Nationality (ISO2)": "AR",
    "Position": "Centre Back",
    "Club": "River Plate",
    "On loan": False,
    "Representative": "Jorge Mendes",
}


class TestXlsxManager(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self.tmp.close()
        os.unlink(self.tmp.name)  # let get_or_create_wb create it fresh

    def tearDown(self):
        if os.path.exists(self.tmp.name):
            os.unlink(self.tmp.name)

    def _load(self):
        wb = openpyxl.load_workbook(self.tmp.name)
        return wb.active

    def test_creates_file_with_headers(self):
        get_or_create_wb(self.tmp.name)
        ws = self._load()
        self.assertEqual(ws.max_row, 1)
        actual = [c.value for c in ws[1]]
        self.assertEqual(actual, HEADERS)

    def test_append_row_count(self):
        append_row(self.tmp.name, PLAYER_1)
        ws = self._load()
        self.assertEqual(ws.max_row, 2)  # header + 1 row

    def test_append_values_player1(self):
        append_row(self.tmp.name, PLAYER_1)
        ws = self._load()
        row = dict(zip(HEADERS, ws.iter_rows(min_row=2, max_row=2, values_only=True).__next__()))
        self.assertEqual(row["Name"], "Matías Abaldo")
        self.assertEqual(row["Birth Date (dd/mm/yyyy)"], "03/04/2004")
        self.assertEqual(row["Height (cm)"], 172)
        self.assertEqual(row["Nationality (ISO2)"], "UY")
        self.assertEqual(row["Club"], "Independiente")
        self.assertEqual(row["On loan"], True)
        self.assertEqual(row["Representative"], "Edgardo Lasalvia")

    def test_append_two_distinct_players(self):
        append_row(self.tmp.name, PLAYER_1)
        append_row(self.tmp.name, PLAYER_2)
        ws = self._load()
        self.assertEqual(ws.max_row, 3)  # header + 2 rows
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        by_header = [dict(zip(HEADERS, r)) for r in rows]
        # Row 2 — player 1
        self.assertEqual(by_header[0]["Name"], "Matías Abaldo")
        self.assertEqual(by_header[0]["Nationality (ISO2)"], "UY")
        # Row 3 — player 2
        self.assertEqual(by_header[1]["Name"], "Lucas Fernández")
        self.assertEqual(by_header[1]["Nationality (ISO2)"], "AR")
        self.assertEqual(by_header[1]["Club"], "River Plate")

    def test_reset_clears_data(self):
        append_row(self.tmp.name, PLAYER_1)
        append_row(self.tmp.name, PLAYER_2)
        reset_rows(self.tmp.name)
        ws = self._load()
        self.assertEqual(ws.max_row, 1)  # only header remains
        self.assertEqual([c.value for c in ws[1]], HEADERS)


if __name__ == "__main__":
    unittest.main(verbosity=2)
