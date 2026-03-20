"""Tests for convert_positions.py"""

import os
import tempfile

import openpyxl
import pytest

from convert_positions import convert_xlsx

SAMPLE_MAP = {
    "Goalkeeper": "PT",
    "Sweeper": "DFC",
    "Centre Back": "DFD",
    "Striker": "DC",
    "Attacking Midfielder": "MP",
    "Right Winger": "ED",
}


def _make_xlsx(rows, headers=("Name", "Position", "Second Position")) -> str:
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    wb.save(tmp.name)
    return tmp.name


def _read_cell(path: str, row: int, col: int):
    return openpyxl.load_workbook(path).active.cell(row, col).value


# ── Basic conversion ──────────────────────────────────────────────────────────

def test_converts_position_column():
    path = _make_xlsx([("Player A", "Goalkeeper", "Striker")])
    try:
        rc = convert_xlsx(path, ["Position", "Second Position"], position_map=SAMPLE_MAP)
        assert rc == 0
        assert _read_cell(path, 2, 2) == "PT"
        assert _read_cell(path, 2, 3) == "DC"
    finally:
        os.unlink(path)


def test_unknown_position_left_unchanged():
    path = _make_xlsx([("Player B", "Unknown Position", None)])
    try:
        convert_xlsx(path, ["Position"], position_map=SAMPLE_MAP)
        assert _read_cell(path, 2, 2) == "Unknown Position"
    finally:
        os.unlink(path)


def test_null_cell_left_unchanged():
    path = _make_xlsx([("Player C", "Goalkeeper", None)])
    try:
        convert_xlsx(path, ["Position", "Second Position"], position_map=SAMPLE_MAP)
        assert _read_cell(path, 2, 2) == "PT"
        assert _read_cell(path, 2, 3) is None
    finally:
        os.unlink(path)


# ── Output file ───────────────────────────────────────────────────────────────

def test_writes_to_separate_output_file():
    path = _make_xlsx([("Player D", "Striker", None)])
    out_path = path + "_out.xlsx"
    try:
        convert_xlsx(path, ["Position"], out_path, position_map=SAMPLE_MAP)
        # Input file is NOT modified
        assert _read_cell(path, 2, 2) == "Striker"
        # Output file has the conversion
        assert _read_cell(out_path, 2, 2) == "DC"
    finally:
        os.unlink(path)
        if os.path.exists(out_path):
            os.unlink(out_path)


def test_overwrites_input_by_default():
    path = _make_xlsx([("Player E", "Goalkeeper", None)])
    try:
        convert_xlsx(path, ["Position"], position_map=SAMPLE_MAP)
        assert _read_cell(path, 2, 2) == "PT"
    finally:
        os.unlink(path)


# ── Multiple rows ─────────────────────────────────────────────────────────────

def test_multiple_rows_all_converted():
    path = _make_xlsx([
        ("Player A", "Goalkeeper", "Striker"),
        ("Player B", "Attacking Midfielder", None),
        ("Player C", "Unknown", "Right Winger"),
    ])
    try:
        convert_xlsx(path, ["Position", "Second Position"], position_map=SAMPLE_MAP)
        assert _read_cell(path, 2, 2) == "PT"
        assert _read_cell(path, 2, 3) == "DC"
        assert _read_cell(path, 3, 2) == "MP"
        assert _read_cell(path, 3, 3) is None
        assert _read_cell(path, 4, 2) == "Unknown"
        assert _read_cell(path, 4, 3) == "ED"
    finally:
        os.unlink(path)


def test_returns_correct_converted_count(capsys):
    path = _make_xlsx([
        ("Player A", "Goalkeeper", "Striker"),
        ("Player B", "Unknown", None),
    ])
    try:
        convert_xlsx(path, ["Position", "Second Position"], position_map=SAMPLE_MAP)
        captured = capsys.readouterr()
        assert "Converted 2 cell(s)" in captured.out
    finally:
        os.unlink(path)


# ── Error handling ────────────────────────────────────────────────────────────

def test_missing_column_returns_error(capsys):
    path = _make_xlsx([("Player A", "Goalkeeper", None)])
    try:
        rc = convert_xlsx(path, ["NonExistentColumn"], position_map=SAMPLE_MAP)
        assert rc == 1
    finally:
        os.unlink(path)


def test_file_not_found_returns_error():
    rc = convert_xlsx("nonexistent_file.xlsx", ["Position"], position_map=SAMPLE_MAP)
    assert rc == 1


def test_partial_valid_columns_proceeds(capsys):
    path = _make_xlsx([("Player A", "Goalkeeper", None)])
    try:
        rc = convert_xlsx(
            path, ["Position", "NonExistentColumn"], position_map=SAMPLE_MAP
        )
        assert rc == 0
        assert _read_cell(path, 2, 2) == "PT"
        captured = capsys.readouterr()
        assert "Warning" in captured.err
    finally:
        os.unlink(path)


# ── Using real position_map.json ──────────────────────────────────────────────

def test_uses_real_map_when_no_map_provided():
    path = _make_xlsx([("Player A", "Goalkeeper", "Striker")])
    try:
        rc = convert_xlsx(path, ["Position", "Second Position"])
        assert rc == 0
        assert _read_cell(path, 2, 2) == "PT"
        assert _read_cell(path, 2, 3) == "DC"
    finally:
        os.unlink(path)


def test_header_row_is_not_modified():
    path = _make_xlsx([("Player A", "Goalkeeper", None)])
    try:
        convert_xlsx(path, ["Position", "Second Position"], position_map=SAMPLE_MAP)
        assert _read_cell(path, 1, 2) == "Position"
        assert _read_cell(path, 1, 3) == "Second Position"
    finally:
        os.unlink(path)
