import os
import shutil
import tempfile
from datetime import datetime

import openpyxl
from openpyxl import Workbook

HEADERS = [
    "Name",
    "Shirt Name",
    "Birth Date (dd/mm/yyyy)",
    "Height (cm)",
    "Weight (kg)",
    "Preferred Foot",
    "Nationality (ISO2)",
    "Second Nationality (ISO2)",
    "Market Value",
    "Position",
    "Second Position",
    "Club",
    "Contract Expiry Date (dd/mm/yyyy)",
    "Notes",
    "Profile Link",
    "Transfermarkt Link",
    "Address",
    "Phone Number",
    "Studies",
    "League",
    "On loan",
    "Club of origin of the Loan",
    "Representative",
    "Representative's Contact",
    "Contract Expiry Date of Representative (dd/mm/yyyy)",
    "Representative Link",
    "Annual Net Salary",
    "Position Profile",
    "Document ID",
    "Document Expiry Date (dd/mm/yyyy)",
]


def get_or_create_wb(path: str) -> openpyxl.Workbook:
    """Open existing xlsx or create one with the correct header row."""
    try:
        wb = openpyxl.load_workbook(path)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(HEADERS)
        _safe_save(wb, path)
    return wb


def append_row(path: str, data: dict) -> str:
    """Append a row to the xlsx. Returns the path actually written to."""
    wb = get_or_create_wb(path)
    ws = wb.active
    row = [data.get(h) or data.get(_short(h)) for h in HEADERS]
    ws.append(row)
    return _safe_save(wb, path)


def reset_rows(path: str) -> None:
    """Delete all data rows below the header, keeping the header intact."""
    wb = get_or_create_wb(path)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    saved = _safe_save(wb, path)
    print(f"Reset: all rows below header deleted from {saved}")


def _safe_save(wb: openpyxl.Workbook, path: str) -> str:
    """
    Write to a temp file first, then atomically replace the target.
    If the target is locked (open in Excel), save to a timestamped
    fallback file instead so no data is lost.
    """
    dir_ = os.path.dirname(os.path.abspath(path))
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=dir_)
    os.close(tmp_fd)

    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, path)
        return path
    except PermissionError:
        # File is locked — save to a dated fallback
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(path)
        fallback = f"{base}_{timestamp}{ext}"
        shutil.move(tmp_path, fallback)
        print(
            f"\n  WARNING: '{os.path.basename(path)}' is open in another program.\n"
            f"  Data saved to fallback: {fallback}\n"
            f"  Close Excel and merge manually, or rerun after closing it."
        )
        return fallback
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _short(header: str) -> str:
    """Return the part of a header before any parenthetical annotation."""
    return header.split("(")[0].strip()
