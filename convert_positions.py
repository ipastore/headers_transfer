#!/usr/bin/env python3
"""
convert_positions.py — Convert position values in an xlsx file using position_map.json.

Usage:
    # Convert Position and Second Position columns in-place (default)
    python convert_positions.py players.xlsx

    # Specify one or more columns
    python convert_positions.py players.xlsx --col "Position"
    python convert_positions.py players.xlsx --col "Position" --col "Second Position"

    # Write to a new file instead of overwriting
    python convert_positions.py players.xlsx --out converted.xlsx

    # Use a custom position map
    python convert_positions.py players.xlsx --map custom_map.json
"""

import argparse
import sys

import openpyxl

from position_mapper import DEFAULT_PLATFORM, convert_position, list_platforms, load_map

DEFAULT_COLUMNS = ["Position", "Second Position"]


def convert_xlsx(
    input_path: str,
    columns: list,
    output_path: str = None,
    position_map: dict = None,
) -> int:
    if position_map is None:
        position_map = load_map()

    try:
        wb = openpyxl.load_workbook(input_path)
    except FileNotFoundError:
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        return 1

    ws = wb.active
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    col_indices = []
    for col in columns:
        if col not in headers:
            print(f"Warning: Column '{col}' not found in headers.", file=sys.stderr)
        else:
            col_indices.append(headers[col])

    if not col_indices:
        print("Error: No valid columns to convert.", file=sys.stderr)
        return 1

    converted = 0
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in col_indices and cell.value:
                original = cell.value
                cell.value = convert_position(str(cell.value), position_map)
                if cell.value != original:
                    converted += 1

    out = output_path or input_path
    wb.save(out)
    print(f"Converted {converted} cell(s). Saved to: {out}")
    return 0


def main():
    parser = argparse.ArgumentParser(
        description="Convert position values in an xlsx file using position_map.json."
    )
    parser.add_argument("input", help="Input .xlsx file path.")
    parser.add_argument(
        "--col",
        metavar="COLUMN",
        action="append",
        dest="columns",
        help=(
            "Column header to convert. Can be repeated. "
            f"Defaults to: {DEFAULT_COLUMNS}"
        ),
    )
    parser.add_argument(
        "--out",
        metavar="OUTPUT",
        help="Output file path. Defaults to overwriting the input file.",
    )
    parser.add_argument(
        "--map",
        metavar="MAP_JSON",
        help="Path to a custom position_map.json (default: docs/position_map.json).",
    )
    parser.add_argument(
        "--platform",
        metavar="PLATFORM",
        default=DEFAULT_PLATFORM,
        help=f"Platform key in the map JSON (default: {DEFAULT_PLATFORM}).",
    )
    args = parser.parse_args()

    columns = args.columns or DEFAULT_COLUMNS
    position_map = load_map(args.map, platform=args.platform)

    sys.exit(convert_xlsx(args.input, columns, args.out, position_map))


if __name__ == "__main__":
    main()
